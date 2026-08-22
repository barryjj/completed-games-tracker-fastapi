"""Spreadsheet import pipeline.

Parses an xlsx file (Google Sheets export), normalises each row, groups rows
that resolve to the same game+platform identity into ImportCandidate records,
and writes them to the DB for user review.

Columns recognised (case-insensitive, order-independent):
  #, Game, Platform, Date, Playthroughs, Notes, Collection

Tabs: one per year; tab name is the fallback year for blank/month-only dates.
"""

import csv
import datetime
import difflib
import re
import time
from io import BytesIO, StringIO
from pathlib import Path

import openpyxl
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from . import match_review, models, titles

# Month name → number (full and abbreviated)
_MONTH_MAP: dict[str, int] = {}
for _i, _names in enumerate(
    [
        ("january", "jan"),
        ("february", "feb"),
        ("march", "mar"),
        ("april", "apr"),
        ("may",),
        ("june", "jun"),
        ("july", "jul"),
        ("august", "aug"),
        ("september", "sep", "sept"),
        ("october", "oct"),
        ("november", "nov"),
        ("december", "dec"),
    ],
    start=1,
):
    for _n in _names:
        _MONTH_MAP[_n] = _i


def _parse_date(raw: str | None, tab_year: int | None) -> tuple[datetime.date | None, str | None]:
    """Normalise a raw date string to a (date, precision) pair.

    precision is 'day' | 'month' | 'year', reflecting what was actually
    knowable from the input — completed_at always holds a full date (1st of
    the month, or Jan 1 for year-only) for sorting purposes, but callers
    that render it to the user should use precision to avoid claiming a
    fabricated day/month is a real one ("January 1, 2012" when the sheet
    only said "2012").

    Accepted formats:
      - Full date: 1/1/2026, 01/01/2026, 2026-01-01           -> day
      - Month + year: "January 2019", "Jan 2019", "1/2019"    -> month
      - Month name only: "January" (uses tab_year)            -> month
      - Blank / None: Jan 1 of tab_year, or None if tab_year
        also unknown                                          -> year
      - Pure year: "2019"                                     -> year
    """
    if not raw:
        if tab_year:
            return datetime.date(tab_year, 1, 1), "year"
        return None, None

    s = str(raw).strip()
    if not s:
        if tab_year:
            return datetime.date(tab_year, 1, 1), "year"
        return None, None

    # openpyxl may hand us a datetime object for formatted cells
    if isinstance(raw, (datetime.date, datetime.datetime)):
        d = raw if isinstance(raw, datetime.date) else raw.date()
        return d, "day"

    # ISO date: 2026-01-01, optionally with a trailing " HH:MM:SS" — the
    # latter shows up when re-parsing ImportRow.raw_date, which stores
    # str(datetime_obj) for cells that were real Excel date values (e.g.
    # "2026-01-01 00:00:00"), not just the date portion.
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})(?: \d{2}:\d{2}:\d{2})?", s)
    if m:
        return datetime.date(int(m[1]), int(m[2]), int(m[3])), "day"

    # Slash full date: 1/1/2026 or 01/01/26
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        y = int(m[3])
        if y < 100:
            y += 2000
        return datetime.date(y, int(m[1]), int(m[2])), "day"

    # Month/year: 1/2019
    m = re.fullmatch(r"(\d{1,2})/(\d{4})", s)
    if m:
        return datetime.date(int(m[2]), int(m[1]), 1), "month"

    # "January 2019" or "Jan 2019"
    m = re.fullmatch(r"([A-Za-z]+)\s+(\d{4})", s)
    if m:
        mon = _MONTH_MAP.get(m[1].lower())
        if mon:
            return datetime.date(int(m[2]), mon, 1), "month"

    # "January" alone — use tab year for the year, but the month itself is
    # genuinely known from the text, so precision is 'month' not 'year'.
    mon = _MONTH_MAP.get(s.lower())
    if mon and tab_year:
        return datetime.date(tab_year, mon, 1), "month"

    # Pure year: "2019"
    m = re.fullmatch(r"(\d{4})", s)
    if m:
        return datetime.date(int(m[1]), 1, 1), "year"

    return None, None


def _parse_playthroughs(raw: str | None) -> str | None:
    """Normalise playthroughs: strip '+', return as string or None."""
    if raw is None:
        return None
    s = str(raw).strip().rstrip("+").strip()
    if not s or s == "0":
        return None
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except ValueError:
        return None


def _tab_year(sheet_title: str) -> int | None:
    """Extract a 4-digit year from a tab name like '2019' or 'Games 2019'."""
    m = re.search(r"\b(20\d{2}|19\d{2})\b", sheet_title)
    return int(m[0]) if m else None


def _col_map(header_row: list) -> dict[str, int]:
    """Return {normalised_name: col_index} from a header row."""
    mapping = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower().lstrip("#").strip()
        if key == "" or key == "#":
            mapping["#"] = i
        else:
            mapping[key] = i
        # Also store raw '#' column
        if str(cell).strip() == "#":
            mapping["#"] = i
    return mapping


def _cell(row: tuple, col_map: dict, *keys: str) -> str | None:
    """Get the first matching key from a row, return stripped string or None."""
    for k in keys:
        idx = col_map.get(k)
        if idx is not None and idx < len(row):
            v = row[idx]
            if v is None:
                continue
            s = str(v).strip() if not isinstance(v, (datetime.date, datetime.datetime)) else v
            if s == "" or s is None:
                continue
            return v if isinstance(v, (datetime.date, datetime.datetime)) else str(v).strip()
    return None


_NUMERAL_RE = re.compile(r"^(?:\d+|[ivxlcdm]+)$")


def _numeral_tokens(normalized_title: str) -> set[str]:
    """Tokens that are digits or Roman numerals — the sequel-identity part
    of a title. Two titles whose numeral tokens differ are different games
    no matter how similar the rest looks ("Golden Axe II" vs "III")."""
    return {t for t in normalized_title.split() if _NUMERAL_RE.fullmatch(t)}


# Number words → digits, so "Episode Two" == "Episode 2" (Steam loves
# spelling them out; spreadsheets love digits). Multi-character Roman
# numerals map too ("Blasphemous II" == "Blasphemous 2") — unambiguous as
# title tokens. Single-letter Roman numerals (I, V, X) are deliberately
# NOT mapped: Mega Man X is not Mega Man 10, I is a pronoun, V is a title.
_WORD_NUMBERS = {
    "zero": "0",
    "one": "1",
    "two": "2",
    "three": "3",
    "four": "4",
    "five": "5",
    "six": "6",
    "seven": "7",
    "eight": "8",
    "nine": "9",
    "ten": "10",
    "eleven": "11",
    "twelve": "12",
    "ii": "2",
    "iii": "3",
    "iv": "4",
    "vi": "6",
    "vii": "7",
    "viii": "8",
    "ix": "9",
    "xi": "11",
    "xii": "12",
    "xiii": "13",
    "xiv": "14",
    "xv": "15",
    "xvi": "16",
    "xvii": "17",
    "xviii": "18",
    "xix": "19",
    "xx": "20",
}


def _normalize_title(title: str) -> str:
    """Fold a title for fuzzy matching and grouping.

    Delegates to titles.normalize_for_match so this path and the PSN merge
    share one implementation — they had drifted, and the PSN copy was deleting
    accented characters rather than folding them (#180).
    """
    return titles.normalize_for_match(title)


def _group_key(title: str, platform_id: int | None, raw_platform: str) -> str:
    """Stable grouping key: normalised title + resolved platform (or raw if unresolved)."""
    t = _normalize_title(title)
    p = str(platform_id) if platform_id is not None else f"raw:{raw_platform.strip().lower()}"
    return f"{t}|{p}"


def _colon_remainder(title: str) -> str | None:
    """Return the text after a colon or " - " subtitle separator, if any."""
    for sep in (":", " - "):
        idx = title.find(sep)
        if idx > 0:
            return title[idx + len(sep) :].strip()
    return None


def _strip_prefix_tokens(raw_title: str, prefix_title: str) -> str | None:
    """Return the tokens remaining after removing `prefix_title`'s normalized
    tokens as a literal prefix of `raw_title`'s, or None if it isn't one.

    Exists because `_colon_remainder` only splits on the FIRST colon/dash in
    raw_title — wrong when the matched base title itself contains a colon
    ("F.E.A.R. 2: Project Origin"). For raw_title "F.E.A.R. 2: Project
    Origin: Reborn", `_colon_remainder` leaves "Project Origin: Reborn" as
    the supposed DLC subtitle, which never matches the DLC child's actual
    name ("F.E.A.R. 2: Reborn (DLC)") since "Project Origin" isn't part of
    it. Stripping the ACTUAL matched entry's full title as the prefix
    (rather than guessing off the first colon) correctly leaves just
    "Reborn".
    """
    raw_toks = match_review._normalise_tokens(raw_title)
    prefix_toks = match_review._normalise_tokens(prefix_title)
    if not prefix_toks or len(raw_toks) <= len(prefix_toks):
        return None
    if raw_toks[: len(prefix_toks)] != prefix_toks:
        return None
    return " ".join(raw_toks[len(prefix_toks) :])


def _title_contains_remainder(remainder: str, candidate_title: str) -> bool:
    """True if `remainder`'s tokens appear as a contiguous, in-order run
    within `candidate_title`'s tokens. DLC titles are typically the full
    base title plus a " - Subtitle" suffix ("The Witcher 3: Wild Hunt -
    Hearts of Stone"), so comparing the bare remainder ("Hearts of Stone")
    against the whole DLC title with the symmetric scorer dilutes badly on
    length — containment is the right check here, not overall similarity.

    Deliberately exact (not fuzzy) token matching: this is a containment
    check, not a statistical score, so match_review's 0.75 fuzzy threshold
    is far too loose here — e.g. "witcher" vs "witches" scores 0.857 and
    would let "The Witcher 2" wrongly containment-match a completely
    unrelated "...Witches and Wizards" title.

    Deliberately ORDERED (substring-of-joined-tokens), not a bag-of-tokens
    check: "Street Fighter II" (-> tokens street/fighter/2, roman numeral
    converted) was wrongly matching "Street Fighter Alpha 2" under a
    bag-of-tokens check, since both titles contain "street", "fighter" and
    a stray "2" — just not adjacent or in the same order. Requiring the
    tokens to appear as a contiguous run fixes this without losing the
    legitimate case (remainder tokens are always meant to appear as a
    literal phrase within the candidate, e.g. "Hearts of Stone" inside
    "...Wild Hunt - Hearts of Stone").
    """
    remainder_toks = match_review._normalise_tokens(remainder)
    candidate_toks = match_review._normalise_tokens(candidate_title)
    if not remainder_toks or not candidate_toks:
        return False
    needle = " ".join(remainder_toks)
    haystack = " ".join(candidate_toks)
    return needle in haystack


def _search_pool(
    db: Session, user_id: int, platform_id: int | None, phrase: str, *, base_only: bool = True
) -> list[models.UserLibraryEntry]:
    """SQL-level narrowing before any Python-side token comparison.

    Pass 1 (normalized, tried first because it's the more reliable match):
    split the phrase into words with the same punctuation-stripping
    `_normalize_title` uses for Python-side comparison, then require every
    word to appear somewhere in the title OR display_name (case-insensitive,
    ANDed across words — narrower than a full scan, but not thrown off by
    stray characters between words). This exists because library titles
    pulled from Steam/etc. often carry decorative symbols the spreadsheet
    text never had (e.g. "Golden Axe™ II" vs spreadsheet "Golden Axe II") —
    a single-string ILIKE on the raw phrase fails outright on the ™, silently
    excluding the correct entry from the pool before any Python normalization
    ever gets a chance to compare them.

    Pass 2 (fallback): the original literal contiguous-phrase ILIKE, kept in
    case the word-split pass is ever too permissive for some phrase shape
    ("that quiet game" splitting into overly common individual words, say) —
    a narrower single-string match is available as a fallback rather than
    always trusting the wider word-AND pool.

    platform_id=None searches across all of the user's platforms — used by
    the collection fallback, where the spreadsheet's platform column often
    doesn't match: a row might say "Steam" (how the collection was played)
    while the specific sub-game's library entry uses its native platform
    (e.g. Castlevania II: Simon's Quest logged as NES, not Steam, even
    though it was played via a Steam collection).
    """
    if not phrase or not phrase.strip():
        return []

    def _base_query():
        q = (
            db.query(models.UserLibraryEntry)
            .join(models.GameRelease, models.UserLibraryEntry.release_id == models.GameRelease.id)
            .join(models.Game, models.GameRelease.game_id == models.Game.id)
            .filter(models.UserLibraryEntry.user_id == user_id)
        )
        if platform_id is not None:
            q = q.filter(models.GameRelease.platform_id == platform_id)
        if base_only:
            q = q.filter(models.Game.parent_id.is_(None), models.Game.is_dlc.is_(False))
        return q

    words = _normalize_title(phrase).split()
    if words:
        q = _base_query()
        for word in words:
            q = q.filter(or_(models.Game.title.ilike(f"%{word}%"), models.Game.display_name.ilike(f"%{word}%")))
        pool = q.all()
        if pool:
            return pool

    q = _base_query().filter(or_(models.Game.title.ilike(f"%{phrase.strip()}%"), models.Game.display_name.ilike(f"%{phrase.strip()}%")))
    return q.all()


def _search_phrase(title: str) -> str:
    """Meaningful substring for SQL-level pool narrowing: the colon/dash
    prefix if the title has a subtitle separator, else the whole title."""
    return match_review._colon_prefix(title) or title


def _collection_match_entry(db: Session, user_id: int, raw_title: str, raw_collection: str) -> models.UserLibraryEntry | None:
    """Fallback tried only when nothing else finds a match: the spreadsheet
    row named a Collection, so use that as a search hint — look for a
    library entry whose title matches raw_collection, then check whether
    any game with parent_id pointing to it matches raw_title. parent_id
    covers both real DLC (is_dlc=True) and a standalone game that simply
    belongs to a collection (is_dlc=False, e.g. ToeJam & Earl under SEGA
    Mega Drive & Genesis Classics) — same field, same meaning either way.

    Deliberately does NOT require Game.is_collection on the matched entry:
    that flag is unreliable in practice (some real collections don't have
    it set) and the point here is just "does this look like a known
    library entry that might explain the row" — a soft hint to try, not a
    strict gate. If nothing matches, the caller falls through to create_new.

    Searches across ALL platforms, not just the row's own platform: a
    spreadsheet row often says how the collection itself was played
    ("Steam"), while a specific sub-game's library entry may use its
    native platform instead (e.g. Castlevania II: Simon's Quest logged as
    NES even though it was played via a Steam
    collection). Restricting to the row's platform would silently miss it.
    """
    pool = _search_pool(db, user_id, None, raw_collection, base_only=True)
    collection_entry = None
    for entry in pool:
        game = entry.release.game if entry.release else None
        if game and _title_contains_remainder(raw_collection, game.title):
            collection_entry = entry
            break
    if not collection_entry:
        return None

    children = (
        db.query(models.UserLibraryEntry)
        .join(models.GameRelease, models.UserLibraryEntry.release_id == models.GameRelease.id)
        .join(models.Game, models.GameRelease.game_id == models.Game.id)
        .filter(
            models.UserLibraryEntry.user_id == user_id,
            models.Game.parent_id == collection_entry.release.game.id,
        )
        .all()
    )
    target = _colon_remainder(raw_title) or raw_title
    for child in children:
        child_title = child.release.game.title if child.release and child.release.game else ""
        if not child_title:
            continue
        if _title_contains_remainder(target, child_title) or _normalize_title(child_title) == _normalize_title(raw_title):
            return child
    return None


# Single-letter Roman numerals and their digit twins — used only by the
# variant fallback below, never mapped unconditionally in normalization.
_SINGLE_NUMERAL_EQUIV = {"i": "1", "v": "5", "x": "10", "1": "i", "5": "v", "10": "x"}


def _single_numeral_variant_entry(db: Session, user_id: int, raw_title: str, platform_id: int) -> models.UserLibraryEntry | None:
    """Sheet says "Final Fantasy X", library has "Final Fantasy 10" (or the
    reverse). Single-letter numerals can't be mapped unconditionally — Mega
    Man X is not Mega Man 10 — but ordering makes the guess safe enough:
    the literal exact pass has already run and found nothing, so try each
    single-token conversion (X↔10, V↔5, I↔1) and require an EXACT hit on
    the variant. The raw titles still differ, so a hit lands flagged as an
    uncertain match and gets human eyes during review."""
    words = _normalize_title(raw_title).split()
    for idx, word in enumerate(words):
        alt = _SINGLE_NUMERAL_EQUIV.get(word)
        if not alt:
            continue
        variant = " ".join(words[:idx] + [alt] + words[idx + 1 :])
        entry = _exact_match_entry(db, user_id, variant, platform_id)
        if entry:
            return entry
    return None


def _widened_pool(db: Session, user_id: int, platform_id: int, raw_title: str, *, base_only: bool) -> list[models.UserLibraryEntry]:
    """_search_pool, retried dropping one word at a time when empty — a
    typo'd or differently-abbreviated word defeats the SQL word filter
    outright, so widen before giving up."""
    pool = list(_search_pool(db, user_id, platform_id, raw_title, base_only=base_only))
    words = _normalize_title(raw_title).split()
    if not pool and 1 < len(words) <= 8:
        seen_ids: set[int] = set()
        for skip in range(len(words)):
            sub = " ".join(words[:skip] + words[skip + 1 :])
            for e in _search_pool(db, user_id, platform_id, sub, base_only=base_only):
                if e.id not in seen_ids:
                    seen_ids.add(e.id)
                    pool.append(e)
    return pool


def _is_subsequence(needle: list[str], hay: list[str]) -> bool:
    it = iter(hay)
    return all(w in it for w in needle)


def _fuzzy_match_entry(db: Session, user_id: int, raw_title: str, platform_id: int) -> models.UserLibraryEntry | None:
    """Near-exact whole-title match for sheet typos ("Fasion Police Squad" →
    "Fashion Police Squad"). Runs after exact, before the structural passes,
    so a one-letter slip doesn't fall through to a prefix match or die in
    create_new. Deliberately strict:
      - similarity >= 0.92 over the full normalized title (difflib)
      - numeral tokens must match exactly (sequels can never fuzzy-match)
      - exactly ONE library entry may clear the bar, else ambiguous → None
      - short titles (< 8 chars spaceless) never fuzzy-match
    A typo'd word also defeats the SQL word filter, so when the pool comes
    back empty the search retries dropping one word at a time."""
    needle = _normalize_title(raw_title)
    if len(needle.replace(" ", "")) < 8:
        return None
    pool = _widened_pool(db, user_id, platform_id, raw_title, base_only=False)
    needle_numerals = _numeral_tokens(needle)
    winners: dict[int, models.UserLibraryEntry] = {}
    for entry in pool:
        game = entry.release.game if entry.release and entry.release.game else None
        if not game:
            continue
        for cand_title in (game.title, game.display_name):
            if not cand_title:
                continue
            norm = _normalize_title(cand_title)
            if _numeral_tokens(norm) != needle_numerals:
                continue
            if difflib.SequenceMatcher(None, needle, norm).ratio() >= 0.92:
                winners[entry.id] = entry
                break
    if len(winners) == 1:
        return next(iter(winners.values()))
    return None


def _prefix_match_entry(db: Session, user_id: int, raw_title: str, platform_id: int) -> models.UserLibraryEntry | None:
    """Structural match, tried before the general fallback: split the
    spreadsheet title on a colon/dash subtitle separator, search the library
    for base games whose title contains the prefix ("The Witcher 3: Hearts
    of Stone" -> search for "The Witcher 3"), then check whether the
    remainder identifies a specific DLC child of that base game ("Hearts of
    Stone") or is otherwise part of the base title itself. Returns None if
    the raw title has no subtitle separator, or no base game's title
    contains the prefix.

    Multiple base games can share a prefix — e.g. "Killing Floor" and
    "Killing Floor: Incursion" are both standalone base games, not DLC. So
    this collects every prefix match first, then prioritizes an exact
    full-title match (pass 1) over a DLC-child match (pass 2). Deliberately
    has NO "just return the first prefix match" fallback — see
    _pool_fallback_entry for why guessing off a bare prefix hit is unsafe.
    """
    prefix = match_review._colon_prefix(raw_title)
    if not prefix:
        return None
    remainder = _colon_remainder(raw_title)

    pool = _search_pool(db, user_id, platform_id, prefix, base_only=True)
    prefix_matched = [
        entry for entry in pool if entry.release and entry.release.game and _title_contains_remainder(prefix, entry.release.game.title)
    ]
    if not prefix_matched:
        return None

    # Pass 1: exact match — remainder (or no remainder) is part of THIS
    # entry's own title, e.g. raw "Killing Floor: Incursion" against base
    # game "Killing Floor: Incursion" itself.
    for entry in prefix_matched:
        game_title = entry.release.game.title
        if not remainder or _title_contains_remainder(remainder, game_title):
            return entry

    # Pass 2: remainder identifies a DLC child of one of the prefix-matched
    # base games (e.g. "Hearts of Stone" under "The Witcher 3: Wild Hunt").
    # Uses each entry's own full title as the prefix to strip (not the naive
    # first-colon split) so base titles that themselves contain a colon
    # ("F.E.A.R. 2: Project Origin") don't leave leftover base-title tokens
    # stuck in front of the real DLC subtitle.
    for entry in prefix_matched:
        entry_remainder = _strip_prefix_tokens(raw_title, entry.release.game.title) or remainder
        if not entry_remainder:
            continue
        child_entries = (
            db.query(models.UserLibraryEntry)
            .join(models.GameRelease, models.UserLibraryEntry.release_id == models.GameRelease.id)
            .join(models.Game, models.GameRelease.game_id == models.Game.id)
            .filter(
                models.UserLibraryEntry.user_id == user_id,
                models.GameRelease.platform_id == platform_id,
                models.Game.parent_id == entry.release.game.id,
            )
            .all()
        )
        for child in child_entries:
            child_title = child.release.game.title if child.release and child.release.game else ""
            if child_title and _title_contains_remainder(entry_remainder, child_title):
                return child

    return None


def _library_prefix_match_entry(db: Session, user_id: int, raw_title: str, platform_id: int) -> models.UserLibraryEntry | None:
    """Reverse of `_prefix_match_entry`: handles spreadsheet titles with no
    subtitle at all ("Sekiro") against a library base game that has one
    ("Sekiro: Shadows Die Twice"). Splits each base game's own title on its
    colon/dash separator and checks for an exact normalized match against
    the raw title — a direct structural comparison, not fuzzy scoring.
    """
    needle = _normalize_title(raw_title)
    pool = _search_pool(db, user_id, platform_id, raw_title, base_only=True)
    for entry in pool:
        game_title = entry.release.game.title if entry.release and entry.release.game else ""
        if not game_title:
            continue
        prefix = match_review._colon_prefix(game_title)
        if prefix and _normalize_title(prefix) == needle:
            return entry
    return None


# Titles containing non-ASCII (ABZÛ, Pokémon...) are invisible to the SQL
# pool search — SQLite LIKE only case-folds ASCII, so '%abzu%' never matches
# ABZÛ and Python-side normalization never gets a candidate to compare.
# This tiny index maps normalized forms of the (few) accented titles to
# their entries. TTL-cached per user+platform so a recheck over hundreds of
# candidates does the library scan once, not per candidate.
_ACCENT_INDEX_TTL = 60.0
_accent_index_cache: dict[tuple[int, int, int], tuple[float, dict[str, int]]] = {}


def _accented_title_index(db: Session, user_id: int, platform_id: int) -> dict[str, int]:
    # Entry count in the key invalidates the cache the moment anything is
    # added (sync mid-recheck, test fixtures, manual adds); the TTL covers
    # renames of existing titles.
    entry_count = (
        db.query(models.UserLibraryEntry.id)
        .join(models.UserLibraryEntry.release)
        .filter(
            models.UserLibraryEntry.user_id == user_id,
            models.GameRelease.platform_id == platform_id,
        )
        .count()
    )
    key = (user_id, platform_id, entry_count)
    cached = _accent_index_cache.get(key)
    now = time.monotonic()
    if cached and now - cached[0] < _ACCENT_INDEX_TTL:
        return cached[1]
    rows = (
        db.query(models.UserLibraryEntry.id, models.Game.title, models.Game.display_name)
        .join(models.UserLibraryEntry.release)
        .join(models.GameRelease.game)
        .filter(
            models.UserLibraryEntry.user_id == user_id,
            models.GameRelease.platform_id == platform_id,
        )
        .all()
    )
    index: dict[str, int] = {}
    for entry_id, title, display_name in rows:
        for t in (title, display_name):
            if not t or t.isascii():
                continue
            norm = _normalize_title(t)
            index.setdefault(norm, entry_id)
            tight = norm.replace(" ", "")
            if len(tight) >= 6:
                index.setdefault(tight, entry_id)
    _accent_index_cache[key] = (now, index)
    return index


def _exact_match_entry(db: Session, user_id: int, raw_title: str, platform_id: int) -> models.UserLibraryEntry | None:
    """Direct normalized-title equality against every entry on the platform
    (base games and DLC alike). Tried before any of the colon-splitting
    heuristics below, since an exact match is the strongest possible signal
    and must never be shadowed by a looser one — e.g. bare "Killing Floor"
    should match the actual "Killing Floor" base game itself, not get
    colon-stripped-matched against the unrelated "Killing Floor: Incursion"
    just because that title happens to start with the same prefix."""
    needle = _normalize_title(raw_title)
    # Spaceless tier: "Blade Chimera" == "BLADECHIMERA". Same-game titles
    # differing only in spacing are effectively exact; distinct games that
    # collide spacelessly would have to collide on every other character
    # too (sequel numerals included), so this can't cross-match sequels.
    # Length floor keeps trivial titles from coincidental collisions.
    needle_tight = needle.replace(" ", "") if len(needle.replace(" ", "")) >= 6 else None
    pool = _search_pool(db, user_id, platform_id, raw_title, base_only=False)
    for entry in pool:
        game = entry.release.game if entry.release and entry.release.game else None
        if not game:
            continue
        # display_name counts too — a user-corrected display name (e.g. the
        # Capcom "2" DLCs renamed to real titles) should be matchable.
        for cand_title in (game.title, game.display_name):
            if not cand_title:
                continue
            norm = _normalize_title(cand_title)
            if norm == needle:
                return entry
            if needle_tight and norm.replace(" ", "") == needle_tight:
                return entry
    # Pool came up empty-handed — check the accented-title index (entries
    # the SQL narrowing structurally cannot find for an ASCII needle).
    accent_index = _accented_title_index(db, user_id, platform_id)
    entry_id = accent_index.get(needle) or (accent_index.get(needle_tight) if needle_tight else None)
    if entry_id:
        return (
            db.query(models.UserLibraryEntry)
            .filter(models.UserLibraryEntry.id == entry_id, models.UserLibraryEntry.user_id == user_id)
            .first()
        )
    return None


def _pool_fallback_entry(db: Session, user_id: int, raw_title: str, platform_id: int) -> models.UserLibraryEntry | None:
    """Last resort when nothing structural confirms a match: narrow the
    candidate pool via the same SQL substring search used above, and only
    accept a match if the pool contains EXACTLY ONE candidate — the SQL
    narrowing itself is then the confirming signal, since nothing else in
    the library even shares the search phrase. If the pool has multiple
    candidates, guessing among them is exactly the mistake this engine was
    rebuilt to avoid: "Contra" narrows to 19 candidates (Contract,
    Contraption Maker, Contrast, three different Contra games...), and none
    of those should get silently picked — return None so it falls to
    create_new instead, where Edit/manual-link can resolve it deliberately.
    """
    phrase = _search_phrase(raw_title)
    # Strict pool for the legacy singleton path below — its whole premise is
    # "the full phrase narrowed the library to exactly one", so it must not
    # see word-drop-widened candidates. The containment tier scans the wide
    # pool: its own guards (in-order subsequence + numeral subset + unique
    # minimum) do the confirming.
    pool_strict = list(_search_pool(db, user_id, platform_id, phrase, base_only=True))
    pool = _widened_pool(db, user_id, platform_id, phrase, base_only=True)

    # Containment tier (works on multi-candidate pools) — deliberately
    # TIGHT after a loose first version mis-matched half a review page:
    #   - forward only: the sheet title inside the candidate, never the
    #     reverse (reverse collapsed every DLC row onto its base game)
    #   - prefix-anchored: the sheet title must be the START of the
    #     candidate, extras only trail ("resident evil 7 biohazard" yes;
    #     "LEGO marvel super heroes" is a different game than "marvel
    #     super heroes", not a decorated one)
    #   - strict numeral equality (bare "Mega Man" must not match "11")
    #   - at most ONE extra token ("resident evil 7 biohazard" yes,
    #     "mega man legacy collection" no)
    #   - strictly unique winner, ties refuse
    needle = _normalize_title(raw_title)
    nwords = needle.split()
    nnums = _numeral_tokens(needle)

    scored: list[tuple[int, models.UserLibraryEntry]] = []
    for cand in pool:
        game = cand.release.game if cand.release and cand.release.game else None
        if not game:
            continue
        best_extras: int | None = None
        for cand_title in (game.title, game.display_name):
            if not cand_title:
                continue
            c = _normalize_title(cand_title)
            cwords = c.split()
            extras = len(cwords) - len(nwords)
            if extras < 0 or extras > 1:
                continue
            if _numeral_tokens(c) != nnums:
                continue
            if cwords[: len(nwords)] != nwords:
                continue
            if best_extras is None or extras < best_extras:
                best_extras = extras
        if best_extras is not None:
            scored.append((best_extras, cand))
    if scored:
        scored.sort(key=lambda t: t[0])
        if len(scored) == 1 or scored[0][0] < scored[1][0]:
            return scored[0][1]
        return None  # tie — genuinely ambiguous, a human must pick

    if len(pool_strict) != 1:
        return None
    entry = pool_strict[0]
    game_title = entry.release.game.title if entry.release and entry.release.game else ""
    if not game_title:
        return None
    # Reverse of the containment rule, same tightness: the LIBRARY title as a
    # contiguous prefix-run of the sheet title with at most one extra sheet
    # token and equal numerals ("resident evil 7 biohazard" in the sheet vs
    # a plain "Resident Evil 7" entry). The old _score >= 0.5 acceptance is
    # gone — it waved "Mega Man 2" into "Mega Man Legacy Collection 2" and
    # was the same rogue class the containment tier's tie rule refuses.
    cand = _normalize_title(game_title)
    if _numeral_tokens(needle) != _numeral_tokens(cand):
        return None
    cwords = cand.split()
    extras = len(nwords) - len(cwords)
    if 0 <= extras <= 1 and nwords[: len(cwords)] == cwords:
        return entry
    return None


def _best_matching_entry(
    db: Session, user_id: int, raw_title: str, platform_id: int | None, raw_collection: str | None = None
) -> models.UserLibraryEntry | None:
    """Find the best existing library entry on the same platform for a
    spreadsheet title. A direct match (the game has its own real library
    entry — whether standalone or DLC anywhere) always wins over anything
    the spreadsheet's Collection column might suggest, since the entry
    already existing is a stronger signal than the row's own metadata.

    Passes in order, each falling through to the next only if it finds
    nothing:
      0. `_exact_match_entry` — direct normalized-title equality, the
         strongest signal, always wins if present.
      1. `_prefix_match_entry` — spreadsheet title has a colon/dash subtitle
         ("The Witcher 3: Hearts of Stone"); split it and search structurally.
      2. `_library_prefix_match_entry` — spreadsheet title has no subtitle
         ("Sekiro") but a library base game does ("Sekiro: Shadows Die
         Twice"); split the library title instead.
      3. `_pool_fallback_entry` — nothing structural confirmed anything;
         accept a single unambiguous candidate from the narrowed pool.
      4. `_collection_match_entry` — only tried if 0-3 all found nothing
         AND the row named a Collection: use it as a search hint to find
         a specific child under that collection (see its docstring).
    """
    if not platform_id:
        return None
    direct = (
        _exact_match_entry(db, user_id, raw_title, platform_id)
        or _single_numeral_variant_entry(db, user_id, raw_title, platform_id)
        or _fuzzy_match_entry(db, user_id, raw_title, platform_id)
        or _prefix_match_entry(db, user_id, raw_title, platform_id)
        or _library_prefix_match_entry(db, user_id, raw_title, platform_id)
        or _pool_fallback_entry(db, user_id, raw_title, platform_id)
    )
    if direct:
        return direct
    if raw_collection and raw_collection.strip():
        return _collection_match_entry(db, user_id, raw_title, raw_collection.strip())
    return None


def rematch_pending_candidates(db: Session, user_id: int) -> int:
    """Re-run title matching against the current library for every pending
    candidate, without re-parsing the spreadsheet. Useful after a sync adds
    a game that was previously unmatched. Returns the number of candidates
    whose match/action changed."""
    candidates = (
        db.query(models.ImportCandidate).filter(models.ImportCandidate.user_id == user_id, models.ImportCandidate.status == "pending").all()
    )
    updated = 0
    for candidate in candidates:
        if not candidate.platform_id:
            continue
        candidate_collection = next((r.raw_collection for r in candidate.rows if r.raw_collection), None)
        best_entry = _best_matching_entry(db, user_id, candidate.raw_title, candidate.platform_id, candidate_collection)
        if best_entry:
            if candidate.library_entry_id != best_entry.id or candidate.proposed_action != "add_to_existing":
                candidate.library_entry_id = best_entry.id
                candidate.proposed_action = "add_to_existing"
                updated += 1
        elif candidate.proposed_action != "create_new":
            candidate.library_entry_id = None
            candidate.proposed_action = "create_new"
            updated += 1
    db.commit()
    return updated


def backfill_completed_at_precision(db: Session, user_id: int) -> tuple[int, int]:
    """One-time repair for rows written before completed_at_precision existed.
    Re-derives precision from ImportRow.raw_date (kept permanently for dedup,
    even for pending rows) and stamps it onto the ImportRow. For rows whose
    candidate is already confirmed, also finds and fixes the Completion it
    produced. Returns (rows_updated, completions_updated).

    Covers both confirmed AND pending candidates — pending rows were parsed
    before this column existed too, so without this they'd silently default
    to 'day' precision whenever they're eventually confirmed.

    Matching a row to its Completion reuses the same key already used for
    row-level dedup on re-import: (completed_at, playthroughs, raw_notes)
    scoped to the candidate's library_entry_id.
    """
    rows = (
        db.query(models.ImportRow)
        .join(models.ImportCandidate, models.ImportRow.candidate_id == models.ImportCandidate.id)
        .filter(
            models.ImportCandidate.user_id == user_id,
            models.ImportCandidate.status.in_(["pending", "confirmed"]),
            models.ImportRow.completed_at_precision.is_(None),
        )
        .options(joinedload(models.ImportRow.candidate))
        .all()
    )
    rows_updated = 0
    completions_updated = 0
    for row in rows:
        if not row.completed_at:
            continue
        tab_year = _tab_year(row.source_tab) if row.source_tab else None
        _, precision = _parse_date(row.raw_date, tab_year)
        if not precision:
            continue
        row.completed_at_precision = precision
        rows_updated += 1

        candidate = row.candidate
        if candidate.status != "confirmed" or not candidate.library_entry_id:
            continue
        completion = (
            db.query(models.Completion)
            .filter(
                models.Completion.library_entry_id == candidate.library_entry_id,
                models.Completion.completed_at == row.completed_at,
                models.Completion.playthroughs == row.playthroughs,
                models.Completion.notes == row.raw_notes,
            )
            .first()
        )
        if completion and completion.completed_at_precision != precision:
            completion.completed_at_precision = precision
            completions_updated += 1
    db.commit()
    return rows_updated, completions_updated


class ParseResult:
    def __init__(self):
        self.candidates: list[dict] = []  # [{raw_title, raw_platform, platform_id, rows:[...]}]
        self.skipped_rows: int = 0
        self.total_rows: int = 0


def _row_values(sheet_row) -> tuple:
    """Convert a row of Cell objects to values, preserving percentage display strings."""
    out = []
    for cell in sheet_row:
        v = cell.value
        if isinstance(v, (int, float)) and cell.number_format and "%" in cell.number_format:
            pct = int(round(v * 100))
            out.append(f"{pct}%")
        else:
            out.append(v)
    return tuple(out)


def _dedup_group_rows(groups: dict) -> None:
    """Drop repeated rows within each group — the same game can appear on
    several tabs, and a re-export can repeat it inside one."""
    for group in groups.values():
        seen: set = set()
        unique_rows = []
        for r in group["rows"]:
            key = (r["completed_at"], r["playthroughs"], r["raw_notes"])
            if key not in seen:
                seen.add(key)
                unique_rows.append(r)
        group["rows"] = unique_rows


# The fields a column can be mapped to. The keys are what _cell() already looks
# up, so a user-supplied mapping and an auto-detected one are the same shape and
# the parser below cannot tell them apart (#197).
IMPORT_FIELDS: tuple[tuple[str, str], ...] = (
    ("game", "Title"),
    ("platform", "Platform"),
    ("date", "Completed date"),
    ("playthroughs", "Playthroughs"),
    ("notes", "Notes"),
    ("collection", "Collection"),
)

# Header text -> field. Everything the parser used to accept, in one place so
# the wizard's guesses and the parser's behaviour cannot drift apart.
_HEADER_GUESSES: dict[str, str] = {
    "game": "game",
    "title": "game",
    "platform": "platform",
    "date": "date",
    "completed": "date",
    "completed date": "date",
    "playthroughs": "playthroughs",
    "times completed": "playthroughs",
    "notes": "notes",
    "collection": "collection",
}


def _find_header_row(rows: list) -> int | None:
    """First row that names a title column. None when nothing looks like one."""
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if any(c in ("game", "title") for c in cells):
            return i
    return None


def _header_labels(rows: list, header_idx: int | None) -> list:
    """The header row with the unnamed leading row-number column filled in.

    Sheets exported from Sheets often have sequential numbers in column A under
    a blank header -- the real file this was built against does exactly that.
    """
    if header_idx is None:
        return []
    header_list = list(rows[header_idx])
    if header_list and header_list[0] is None and "#" not in _col_map(header_list):
        header_list[0] = "#"
    return header_list


def _guess_columns(header_list: list) -> dict[str, int]:
    """Auto-detected {field: column index} — what the parser did on its own."""
    cols: dict[str, int] = {}
    for i, cell in enumerate(header_list):
        if cell is None:
            continue
        field = _HEADER_GUESSES.get(str(cell).strip().lower().lstrip("#").strip() or "#")
        if field and field not in cols:
            cols[field] = i
    return cols


def _parse_sheet(
    rows: list,
    sheet_title: str,
    groups: dict,
    result: "ParseResult",
    db: Session,
    user_id: int,
    spec: dict | None = None,
) -> None:
    """Fold one sheet's worth of rows into `groups`.

    Split out of parse_xlsx so a CSV can use it too. A CSV is one sheet with no
    tabs, and nothing below this line cares where the rows came from -- rows are
    rows (#197).

    `sheet_title` is only a fallback for the year: a row whose Date column
    carries a full date already knows its year, and _parse_date only reaches for
    the tab year when the date is a bare month or blank. So a single sheet with
    complete dates behaves identically whether it arrived as xlsx or csv.
    """
    if not rows:
        return

    # `spec` is the wizard's answer for this sheet: which row is the header,
    # which column is which field, and what year to fall back on. Without one
    # the sheet is auto-detected exactly as before, so every existing caller is
    # unchanged (#197).
    if spec is None:
        header_idx = _find_header_row(rows)
        if header_idx is None:
            result.skipped_rows += len(rows)
            return
        cols = _guess_columns(_header_labels(rows, header_idx))
        tab_year = _tab_year(sheet_title)
    else:
        header_idx = spec.get("header_row")
        cols = {f: i for f, i in (spec.get("cols") or {}).items() if i is not None}
        tab_year = spec.get("year") or _tab_year(sheet_title)
        if "game" not in cols:
            # Nothing to import without a title, and silently skipping is the
            # behaviour this whole feature exists to remove.
            result.skipped_rows += len(rows)
            return
        # A headerless sheet starts at row 0; -1 makes the slice below work out.
        if header_idx is None:
            header_idx = -1

    # Position IS the row number. It exists only to become Completion.sort_order,
    # which is what keeps two completions in the same month in the order the
    # sheet had them -- month-precision dates all land on the 1st, so without it
    # they have nothing to order by. Reading it out of a "#" column asked the
    # user to map a column whose only job was to restate the order the rows were
    # already in, and it did not work at all on a sheet that has no such column
    # (#197).
    position = 0
    for row in rows[header_idx + 1 :]:
        result.total_rows += 1

        # Skip entirely blank rows
        if all(c is None or str(c).strip() == "" for c in row):
            result.skipped_rows += 1
            continue

        raw_title = _cell(row, cols, "game", "title")
        if not raw_title:
            result.skipped_rows += 1
            continue

        raw_platform = _cell(row, cols, "platform") or ""
        raw_date = _cell(row, cols, "date")
        raw_playthroughs = _cell(row, cols, "playthroughs", "times completed")
        raw_notes = _cell(row, cols, "notes")
        raw_collection = _cell(row, cols, "collection")

        position += 1
        row_number = position

        platform_str = re.split(r"[·|/]", raw_platform)[0].strip() if raw_platform else ""
        platform_id = models.resolve_platform_id(db, platform_str) if platform_str else None
        completed_at, completed_at_precision = _parse_date(raw_date, tab_year)
        playthroughs = _parse_playthroughs(raw_playthroughs)

        key = _group_key(raw_title, platform_id, raw_platform)

        if key not in groups:
            groups[key] = {
                "raw_title": raw_title,
                "raw_platform": raw_platform,
                "platform_id": platform_id,
                "rows": [],
            }

        groups[key]["rows"].append(
            {
                "raw_title": raw_title,
                "raw_platform": raw_platform,
                "raw_date": str(raw_date) if raw_date else None,
                "raw_playthroughs": str(raw_playthroughs) if raw_playthroughs else None,
                "raw_notes": raw_notes,
                "raw_collection": raw_collection,
                "source_tab": sheet_title,
                "row_number": row_number,
                "completed_at": completed_at,
                "completed_at_precision": completed_at_precision,
                "playthroughs": playthroughs,
            }
        )


def _sheets_from_upload(file_bytes: bytes, filename: str) -> list[tuple[str, list]]:
    """(sheet name, rows) for either format. One entry for a CSV, named after
    the file, since the name is where a CSV keeps its year."""
    if filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        rows = [[(c if c.strip() != "" else None) for c in row] for row in csv.reader(StringIO(text), dialect)]
        return [(Path(filename).stem, rows)]
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    return [(sheet.title, [_row_values(r) for r in sheet.iter_rows()]) for sheet in wb.worksheets]


def inspect_upload(file_bytes: bytes, filename: str, db: Session, samples: int = 3) -> list[dict]:
    """What the parser THINKS it found, without importing anything.

    The importer used to guess in silence: a sheet whose header it did not
    recognise was skipped whole, and a column named "Completed" instead of
    "Date" dropped every date without a word. You found out days later, by
    noticing the count was wrong.

    This returns the guesses so they can be shown and corrected first. Per
    sheet: the header row, every column with the field it would be read as, and
    a few rows rendered the way they would actually be imported -- the parsed
    date with its precision, the resolved platform, the parsed playthrough
    count. Nothing is written and no candidates are built (#197).
    """
    out = []
    for name, rows in _sheets_from_upload(file_bytes, filename):
        header_idx = _find_header_row(rows)
        labels = _header_labels(rows, header_idx)
        cols = _guess_columns(labels)
        width = max((len(r) for r in rows[: (header_idx or 0) + 1 + samples]), default=0)
        by_index = {i: f for f, i in cols.items()}
        columns = [
            {
                "index": i,
                "letter": _column_letter(i),
                "label": (str(labels[i]).strip() if header_idx is not None and i < len(labels) and labels[i] is not None else ""),
                "field": by_index.get(i, ""),
            }
            for i in range(width)
        ]
        year = _tab_year(name)
        data = [r for r in rows[(header_idx + 1) if header_idx is not None else 0 :] if any(c is not None and str(c).strip() for c in r)]
        # A row with no title is skipped by the parser. Usually that is a
        # pre-numbered empty row waiting to be filled in -- the real file has 20
        # of them -- but "52 rows" with no further explanation invites you to
        # wonder where the other 20 went, so both numbers are reported.
        importable = [r for r in data if (_cell(r, cols, "game", "title") or "").strip()]
        preview = []
        for row in importable[:samples]:
            raw_date = _cell(row, cols, "date")
            raw_playthroughs = _cell(row, cols, "playthroughs", "times completed")
            completed, precision = _parse_date(raw_date, year)
            raw_platform = _cell(row, cols, "platform") or ""
            platform_str = re.split(r"[·|/]", raw_platform)[0].strip() if raw_platform else ""
            preview.append(
                {
                    "title": _cell(row, cols, "game", "title"),
                    "platform": raw_platform,
                    # The resolved row, so the preview can show the same badge
                    # the review list shows rather than a bare word. None means
                    # unmatched, which the template flags the way an unlinked
                    # import row does.
                    "platform_row": (db.get(models.Platform, models.resolve_platform_id(db, platform_str)) if platform_str else None),
                    "completed": completed,
                    "precision": precision,
                    "playthroughs": _parse_playthroughs(raw_playthroughs),
                    "raw_playthroughs": raw_playthroughs,
                    # Marked in the preview and explained once underneath,
                    # rather than repeating "from 2+" on every affected row.
                    "playthroughs_coerced": bool(raw_playthroughs) and str(raw_playthroughs) != str(_parse_playthroughs(raw_playthroughs)),
                    "date_vague": bool(precision) and precision != "day",
                    "collection": _cell(row, cols, "collection"),
                }
            )
        # Does this sheet actually need a fallback year? Only if some row's date
        # cannot supply one on its own -- a blank cell, or a bare month like
        # "January". Scanned across every importable row, not just the three
        # shown, so a file with complete dates is never asked a question it has
        # already answered.
        needs_year = any(_parse_date(_cell(r, cols, "date"), None)[0] is None for r in importable)

        # One note per kind, under the table, instead of a parenthetical on
        # every row that happens to trip it.
        notes = []
        if any(p["playthroughs_coerced"] for p in preview):
            notes.append(("Playthroughs", "must be a whole number — anything else in the cell is dropped."))
        if any(p["date_vague"] for p in preview):
            notes.append(("Completed", "no day given, so the completion is recorded to the month or year only."))
        out.append(
            {
                "name": name,
                "notes": notes,
                "needs_year": needs_year,
                "header_row": header_idx,
                "columns": columns,
                "cols": cols,
                "year": year,
                "data_rows": len(data),
                "importable_rows": len(importable),
                "preview": preview,
            }
        )
    return out


def _column_letter(i: int) -> str:
    """0 -> A, 25 -> Z, 26 -> AA. For naming columns on a headerless sheet."""
    out = ""
    i += 1
    while i:
        i, rem = divmod(i - 1, 26)
        out = chr(65 + rem) + out
    return out


def parse_csv(file_bytes: bytes, filename: str, db: Session, user_id: int) -> ParseResult:
    """Parse a CSV export. One sheet, named after the file.

    A CSV is a spreadsheet with the tabs taken away, and the tabs were only ever
    a source of the YEAR -- and only for rows whose Date column does not carry
    one. Everything else about parsing is row-level, so a CSV goes through the
    same code as a worksheet (#197).

    The filename stands in for the tab name, which means "Completed Games -
    2026.csv" still dates a row that only says "January". A file with no year in
    its name degrades exactly the way a badly-named tab does today: rows with
    complete dates are fine, rows without one lose their year.
    """
    text = file_bytes.decode("utf-8-sig", errors="replace")
    # Sniff the delimiter -- exports are comma or tab depending on where the
    # Export as menu was clicked. Falls back to comma on an unreadable sample.
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    rows = [[(c if c.strip() != "" else None) for c in row] for row in csv.reader(StringIO(text), dialect)]

    result = ParseResult()
    groups: dict[str, dict] = {}
    _parse_sheet(rows, Path(filename).stem, groups, result, db, user_id)
    _dedup_group_rows(groups)
    result.candidates = list(groups.values())
    return result


def parse_upload(file_bytes: bytes, filename: str, db: Session, user_id: int, specs: dict | None = None) -> ParseResult:
    """Parse an uploaded file of either format. The one entry point the job uses.

    `specs` is the wizard's answers, keyed by sheet name: which row is the
    header, which column is which field, which year to fall back on. Without it
    every sheet is auto-detected exactly as before.
    """
    result = ParseResult()
    groups: dict[str, dict] = {}
    for name, rows in _sheets_from_upload(file_bytes, filename):
        _parse_sheet(rows, name, groups, result, db, user_id, spec=(specs or {}).get(name))
    _dedup_group_rows(groups)
    result.candidates = list(groups.values())
    return result


def parse_xlsx(file_bytes: bytes, db: Session, user_id: int) -> ParseResult:
    """Parse an xlsx file and return grouped ImportCandidate data (not yet written to DB)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    result = ParseResult()

    # groups: group_key → {raw_title, raw_platform, platform_id, rows:[row_dict,...]}
    groups: dict[str, dict] = {}

    for sheet in wb.worksheets:
        _parse_sheet([_row_values(r) for r in sheet.iter_rows()], sheet.title, groups, result, db, user_id)

    _dedup_group_rows(groups)
    result.candidates = list(groups.values())
    return result


_BATCH_SIZE = 25


def write_candidates(result: ParseResult, db: Session, user_id: int, on_progress=None) -> int:
    """Write parsed groups to ImportCandidate + ImportRow rows in small batches. Returns candidate count."""

    count = 0
    skipped = 0
    for group in result.candidates:
        # Skip groups that already exist (pending or confirmed) from a previous upload
        plat_filter = (
            models.ImportCandidate.platform_id == group["platform_id"]
            if group["platform_id"] is not None
            else models.ImportCandidate.platform_id.is_(None)
        )
        # Skip if already staged (pending) — don't create a duplicate in the same session
        already_pending = (
            db.query(models.ImportCandidate)
            .filter(
                models.ImportCandidate.user_id == user_id,
                models.ImportCandidate.raw_title == group["raw_title"],
                plat_filter,
                models.ImportCandidate.status == "pending",
            )
            .first()
        )
        if already_pending:
            skipped += 1
            continue

        # Filter out individual rows already confirmed in a previous import.
        # Checks ALL confirmed candidates for this title+platform, not just one —
        # multiple confirmed candidates can accumulate for the same title across
        # separate uploads over time (e.g. DLC rows added in later sessions).
        confirmed_row_keys: set[tuple] = set()
        confirmed_candidates = (
            db.query(models.ImportCandidate)
            .filter(
                models.ImportCandidate.user_id == user_id,
                models.ImportCandidate.raw_title == group["raw_title"],
                plat_filter,
                models.ImportCandidate.status == "confirmed",
            )
            .all()
        )
        for confirmed_candidate in confirmed_candidates:
            for r in confirmed_candidate.rows:
                confirmed_row_keys.add((r.completed_at, r.playthroughs, r.raw_notes))

        new_rows = [r for r in group["rows"] if (r["completed_at"], r["playthroughs"], r["raw_notes"]) not in confirmed_row_keys]
        if not new_rows:
            skipped += 1
            continue
        group["rows"] = new_rows

        # Look for an existing library entry matching title + platform
        group_collection = next((r["raw_collection"] for r in group["rows"] if r.get("raw_collection")), None)
        existing_entry = _best_matching_entry(db, user_id, group["raw_title"], group["platform_id"], group_collection)

        if existing_entry:
            action = "add_to_existing"
        elif group["platform_id"] is None:
            action = "needs_review"
        else:
            action = "create_new"

        candidate = models.ImportCandidate(
            user_id=user_id,
            raw_title=group["raw_title"],
            raw_platform=group["raw_platform"],
            platform_id=group["platform_id"],
            library_entry_id=existing_entry.id if existing_entry else None,
            status="pending",
            proposed_action=action,
        )
        db.add(candidate)
        db.flush()

        for row in group["rows"]:
            db.add(
                models.ImportRow(
                    candidate_id=candidate.id,
                    raw_title=row["raw_title"],
                    raw_platform=row["raw_platform"],
                    raw_date=row["raw_date"],
                    raw_playthroughs=row["raw_playthroughs"],
                    raw_notes=row["raw_notes"],
                    raw_collection=row["raw_collection"],
                    source_tab=row["source_tab"],
                    row_number=row["row_number"],
                    completed_at=row["completed_at"],
                    completed_at_precision=row["completed_at_precision"],
                    playthroughs=row["playthroughs"],
                )
            )
        count += 1
        if count % _BATCH_SIZE == 0:
            db.commit()
        if on_progress:
            on_progress(count)

    if count % _BATCH_SIZE != 0:
        db.commit()
    return count
